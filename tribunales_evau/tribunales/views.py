import logging
from math import ceil

from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.cache import cache
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from django.shortcuts import render
from django.urls import reverse_lazy
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font
from pulp import LpMinimize, LpProblem, LpVariable, lpSum, value
from pulp.apis import PULP_CBC_CMD
from tribunales.models import Asignatura, Evaluador, Examen, Sede

logger = logging.getLogger(__name__)

# CACHE_PREFIX = "moves_"

weekday_names_es = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

M = 10000


def get_fechas(asignatura):
    exams = Examen.objects.filter(COD_ASIGNATURA=asignatura)
    fechas = exams.values_list("FECHA", flat=True).distinct()
    return fechas


def get_hqs(asignatura, fecha):
    sedes = Sede.objects.all()

    headquarter_data = {}
    for sede in sedes:
        sede_id = sede.COD_SEDE
        try:
            examenes = Examen.objects.get(COD_SEDE=sede_id, COD_ASIGNATURA=asignatura, FECHA=fecha)
            exams = examenes.EXAMENES
        except ObjectDoesNotExist:
            logger.debug(f"No Examen for sede={sede_id}, asignatura={asignatura} and fecha={fecha}")
            exams = 0
        try:
            evaluadores = Evaluador.objects.get(COD_SEDE=sede_id, COD_ASIGNATURA=asignatura)
            evals = evaluadores.EVALUADORES
        except ObjectDoesNotExist:
            logger.debug(f"No Evaluador for sede={sede_id} and asignatura={asignatura}")
            evals = 0
        if str(sede_id) not in headquarter_data:
            headquarter_data[str(sede_id)] = {"exams": 0, "evals": 0}
        headquarter_data[str(sede_id)]["exams"] += exams
        headquarter_data[str(sede_id)]["evals"] += evals

    logger.debug("hq_data: " + str(headquarter_data))

    return headquarter_data


def problem_solve(headquarter_data, mean, HQs_from, HQs_to):
    # LP problem
    prob = LpProblem("Headquarters_Movement", LpMinimize)

    # num_exchanges = LpVariable("num_exchanges", lowBound=0, cat="Integer")  # The number of exchanges

    moves = LpVariable.dicts("moves", (HQs_from, HQs_to), lowBound=0, cat="Integer")  # number of moved exams
    exchanges = LpVariable.dicts("is_moved", (HQs_from, HQs_to), cat="Binary")

    # Total number of exams moved -> not needed
    # prob += num_moves == lpSum(moves[i][j] for i in HQ for j in HQ)

    # Objective: Minimize total number of exchanges
    prob += lpSum(exchanges[i][j] for i in HQs_from for j in HQs_to)

    # Constraints
    for i in HQs_from:
        for j in HQs_to:
            prob += exchanges[i][j] <= moves[i][j]
            prob += exchanges[i][j] * M >= moves[i][j]  # M is a large number
            prob += moves[i][j] >= 0
            prob += moves[i][j] <= max(
                0,
                min(
                    headquarter_data[i]["exams"] - (mean * headquarter_data[i]["evals"]),  # los que sobran a i
                    max(
                        (mean * headquarter_data[j]["evals"]) - headquarter_data[j]["exams"], 0
                    ),  # los que le faltan a j
                ),
            )

    for i in HQs_from:
        prob += lpSum(moves[i][j] for j in HQs_to) == max(
            headquarter_data[i]["exams"] - (mean * headquarter_data[i]["evals"]), 0
        )

    for j in HQs_to:
        prob += (
            lpSum(moves[i][j] for i in HQs_from)
            == (mean * headquarter_data[j]["evals"]) - headquarter_data[j]["exams"]
        )

    # prob.solve(PULP_CBC_CMD(msg=False, options=['TimeLimit=60'], mip=True)) # Time-limited version
    prob.solve(PULP_CBC_CMD(msg=False, mip=True))

    print("moves")
    for i in HQs_from:
        print([moves[i][j].value() for j in HQs_to])
    print("exchanges")
    for i in HQs_from:
        print([exchanges[i][j].value() for j in HQs_to])
    print(f"obj_value: {prob.objective.value()}")

    move_details = {}
    for HQ_from in moves:
        for HQ_to in moves[HQ_from]:
            if int(value(moves[HQ_from][HQ_to])) > 0:
                if HQ_from not in move_details:
                    move_details[HQ_from] = {}
                move_details[HQ_from][HQ_to] = int(value(moves[HQ_from][HQ_to]))
                # print(str(HQ_from)+","+str(HQ_to)+": "+str(value(moves[HQ_from][HQ_to]))+
                #       "*"+str(value(exchanges[HQ_from][HQ_to]))+" = "+str(move_details[HQ_from][HQ_to]))

    return move_details


def get_moves(asignatura, fecha):
    move_data = cache.get(str(asignatura.COD_ASIGNATURA) + "_" + str(fecha))

    if move_data is not None:
        return move_data

    logger.debug("Recalculating moves for " + str(asignatura))

    headquarter_data = get_hqs(asignatura, fecha)

    if sum(data["evals"] for data in headquarter_data.values()) == 0:
        logger.debug("No data in DB")
        return {"mean": None, "total_moves": None, "move_details": None}

    mean = ceil(
        # mean = floor(
        sum(data["exams"] for data in headquarter_data.values())
        / sum(data["evals"] for data in headquarter_data.values())
    )

    HQs = list(headquarter_data.keys())
    HQs_from = []
    HQs_to = []
    for HQ in HQs:
        if mean * headquarter_data[HQ]["evals"] > headquarter_data[HQ]["exams"]:
            HQs_to.append(HQ)
        else:
            HQs_from.append(HQ)

    move_details = problem_solve(headquarter_data, mean, HQs_from, HQs_to)

    total_moves = 0
    for key in move_details:
        total_moves += sum(move_details[key].values())

    move_data = {"mean": mean, "total_moves": total_moves, "move_details": move_details}

    cache.set(str(asignatura.COD_ASIGNATURA) + "_" + str(fecha), move_data, timeout=None)
    return cache.get(str(asignatura.COD_ASIGNATURA) + "_" + str(fecha))


class MovesView(LoginRequiredMixin, View):
    login_url = reverse_lazy("account_login")

    def get(self, request):
        cod_asignatura_fecha = request.GET.get("asignatura")

        asignaturas = Asignatura.objects.all()
        self.asignaturas_fecha = []
        for asignatura in asignaturas:
            fechas = get_fechas(asignatura)
            for fecha in fechas:
                self.asignaturas_fecha.append(
                    (str(asignatura.COD_ASIGNATURA) + "_" + str(fecha), f"{asignatura.ASIGNATURA} ({fecha})")
                )

        # When called after form was completed, you'll have cod_asignatura filled
        if cod_asignatura_fecha is not None:
            parts = cod_asignatura_fecha.split("_")
            cod_asignatura = parts[0]
            fecha = parts[1]
            asignatura = Asignatura.objects.get(COD_ASIGNATURA=cod_asignatura)
            self.nombre_asignatura = asignatura.ASIGNATURA + (f" ({fecha})")

            move_data = get_moves(asignatura, fecha)
            move_details = []
            if move_data["move_details"] is not None:
                for from_HQ in move_data["move_details"].keys():
                    for to_HQ in move_data["move_details"][from_HQ].keys():
                        from_sede = Sede.objects.get(COD_SEDE=from_HQ).UBICACION
                        to_sede = Sede.objects.get(COD_SEDE=to_HQ).UBICACION
                        n_exams = int(move_data["move_details"][from_HQ][to_HQ])
                        if n_exams > 0:
                            move_details.append(
                                {
                                    "from_HQ": from_sede,
                                    "to_HQ": to_sede,
                                    "num_moves": n_exams,
                                }
                            )

            return render(
                request,
                "moves_template.html",
                {
                    "asignaturas": self.asignaturas_fecha,
                    "nombre_asignatura": self.nombre_asignatura,
                    "mean": move_data["mean"],
                    "total_moves": move_data["total_moves"],
                    "move_details": move_details,
                },
            )
        else:
            return render(request, "moves_template.html", {"asignaturas": self.asignaturas_fecha})


class MovesXLSView(LoginRequiredMixin, View):
    login_url = reverse_lazy("account_login")

    moves_data = {}

    def print_header(self, ws, sede):
        sede_title = [" ", "SEDE " + str(sede), " ", " ", "Intercambio de exámenes con otras sedes"]
        ws.append(sede_title)
        row = ws.row_dimensions[1]
        row.font = Font(bold=True)

        ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=8)
        ws.merge_cells(start_row=3, start_column=9, end_row=3, end_column=11)
        ws["F3"] = "# Exámenes a Enviar"
        ws["I3"] = "# Exámenes a recibir"

        header = [
            " ",
            "ASIGNATURA",
            "Nº Exám/Sede",
            "nº Correctores/sede",
            "MEDIA UAM",
            "Teóricos",
            "Reales",
            "Sede\ndestino",
            "Teóricos",
            "Reales",
            "Sede\norigen",
        ]
        ws.append(header)

    def get(self, request):
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = 'attachment; filename="sede_data.xlsx"'

        wb = Workbook()
        # Remove the default first sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Fetch data from DB
        sedes = Sede.objects.all()

        for sede in sedes:
            # Create a new sheet for each SEDE
            ws = wb.create_sheet(title="SEDE " + str(sede.COD_SEDE))
            self.print_header(ws, sede.COD_SEDE)

            queryset = (
                Examen.objects.filter(COD_SEDE_id=sede.COD_SEDE)
                .select_related("COD_ASIGNATURA")
                .order_by("FECHA", "COD_ASIGNATURA")
            )

            # Write data rows
            for examen in queryset:
                try:
                    evaluadores = Evaluador.objects.get(
                        COD_SEDE=sede.COD_SEDE, COD_ASIGNATURA=examen.COD_ASIGNATURA
                    ).EVALUADORES
                except ObjectDoesNotExist:
                    evaluadores = 0

                # get_moves
                if examen.COD_ASIGNATURA not in self.moves_data:
                    asignatura = Asignatura.objects.get(COD_ASIGNATURA=examen.COD_ASIGNATURA_id)
                    self.moves_data[examen.COD_ASIGNATURA_id] = get_moves(asignatura, examen.FECHA)
                moves = self.moves_data[examen.COD_ASIGNATURA_id]

                HQ = str(examen.COD_SEDE_id)
                if "move_details" in moves and moves["move_details"] is not None:
                    if HQ in moves["move_details"]:
                        for to_HQ in moves["move_details"][HQ].keys():
                            if moves["move_details"][HQ][to_HQ] > 5:
                                row_data = [
                                    weekday_names_es[examen.FECHA.isoweekday() - 1],
                                    examen.COD_ASIGNATURA.ASIGNATURA,
                                    examen.EXAMENES,
                                    evaluadores,
                                    moves["mean"],  # media_uam,
                                    moves["move_details"][HQ][to_HQ],  # envio_teoricos,
                                    " ",  # envio_reales,
                                    to_HQ,  # envio_sede_destino,
                                    " ",  # recibido_teoricos,
                                    " ",  # recibido_reales,
                                    " ",  # asignatura.recibido_sede_origen
                                ]
                                ws.append(row_data)
                    else:
                        for from_HQ in moves["move_details"].keys():
                            if HQ in moves["move_details"][from_HQ].keys() and moves["move_details"][from_HQ][HQ] > 5:
                                row_data = [
                                    weekday_names_es[examen.FECHA.isoweekday() - 1],
                                    examen.COD_ASIGNATURA.ASIGNATURA,
                                    examen.EXAMENES,
                                    evaluadores,
                                    moves["mean"],  # media_uam,
                                    " ",  # envio_teoricos,
                                    " ",  # envio_reales,
                                    " ",  # envio_sede_destino,
                                    moves["move_details"][from_HQ][HQ],  # recibido_teoricos,
                                    " ",  # recibido_reales,
                                    from_HQ,  # asignatura.recibido_sede_origen
                                ]
                                ws.append(row_data)
        # Save the workbook to the response
        logger.debug("saving workbook")
        wb.save(response)

        return response
